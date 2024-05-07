from openpyxl import load_workbook, Workbook



def fetch_workbook(filepath):
    wb = load_workbook(filepath)
    return wb

def fetch_worksheet(wb, sheet_name):
    ws = wb[sheet_name]
    return ws

day_2_survey_data_file = '../support_files/day_2_house_ids.xlsx'
day_3_survey_data_file = '../support_files/day_3_house_ids.xlsx'
day_4_survey_data_file = '../support_files/day_4_house_ids.xlsx'


def get_house_id_and_index(filepath, sheet_name):
    house_id_and_index = {}
    workbook = fetch_workbook(filepath)
    worksheet = fetch_worksheet(workbook, sheet_name)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2, values_only=True):
        house_id_and_index[row[1]] = row[0] 

    return house_id_and_index

def get_parent_indexes(filepath, sheet_name):
    indexes = []
    workbook = fetch_workbook(filepath)
    worksheet = fetch_worksheet(workbook, sheet_name)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True):
        indexes.append(row[0])

    return indexes


def merge_index_to_house_id(filepath):
    house_id_and_index = get_house_id_and_index(filepath, 'main_survey')
    parent_indexes = get_parent_indexes(filepath, 'cont_survey')
    single_house_id_and_index = tuple()
    house_id_and_index_list = []
    for index in parent_indexes:
        single_house_id_and_index = (index, str(house_id_and_index[index]).replace(" ", ""))
        house_id_and_index_list.append(single_house_id_and_index)

    return house_id_and_index_list


def write_to_file(source_filepath, output_file):
    house_id_and_index_list = merge_index_to_house_id(source_filepath)
    wb = Workbook()
    ws = wb.active
    ws.append(['Index', 'House ID'])
    for house_id_and_index in house_id_and_index_list:
        ws.append(house_id_and_index)
    
    wb.save(f'{output_file}.xlsx')

    print('File has been written successfully')
    return True


write_to_file(day_4_survey_data_file, 'day_4_ids')
write_to_file(day_3_survey_data_file, 'day_3_ids')
write_to_file(day_2_survey_data_file, 'day_2_ids')



    

# print(merge_index_to_house_id(day_4_survey_data_file))
# print(get_house_id_and_index('../support_files/day_2_house_ids.xlsx', 'main_survey'))
# print(get_parent_indexes(day_4_survey_data_file, 'cont_survey'))



   
    





